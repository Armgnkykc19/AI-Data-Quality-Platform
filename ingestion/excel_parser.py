from __future__ import annotations

from pathlib import Path

from ingestion.config import IngestionConfig
from ingestion.errors import (
    BlankHeaderCell,
    ColumnLimitExceeded,
    DuplicateHeader,
    EmptyFile,
    MissingHeader,
    RowLimitExceeded,
    WorkbookError,
    WorksheetNotFound,
)
from ingestion.models import ParsedDataset, ParsedRow, ParseIssue, RejectedRow, SourceMetadata


def _cell_to_str(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _is_entirely_blank(values: list[str | None]) -> bool:
    return all(value is None for value in values)


def parse_xlsx_file(
    path: Path,
    config: IngestionConfig,
    worksheet_name: str | None = None,
) -> ParsedDataset:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise WorkbookError(
            code="missing_dependency",
            message="openpyxl is required for XLSX ingestion.",
        ) from exc

    size = path.stat().st_size
    if size == 0:
        raise EmptyFile(code="empty_file", message="Input workbook is empty.")
    if size > config.max_file_size_bytes:
        from ingestion.errors import FileTooLarge

        raise FileTooLarge(
            code="file_too_large",
            message=f"File size {size} exceeds limit {config.max_file_size_bytes}.",
            size_bytes=size,
            limit_bytes=config.max_file_size_bytes,
        )

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise WorkbookError(
            code="workbook_error",
            message=f"Unable to open workbook: {exc}",
        ) from exc

    available = tuple(workbook.sheetnames)
    if not available:
        raise WorkbookError(code="workbook_error", message="Workbook contains no worksheets.")

    selected = worksheet_name
    if selected is None:
        if config.excel_worksheet_selection == "first":
            selected = available[0]
        else:
            raise WorksheetNotFound(
                code="worksheet_not_found",
                message="Explicit worksheet selection is required by configuration.",
                worksheet="",
            )
    elif selected not in available:
        raise WorksheetNotFound(
            code="worksheet_not_found",
            message=f"Worksheet not found: {selected}",
            worksheet=selected,
        )

    sheet = workbook[selected]
    metadata = SourceMetadata(
        path=str(path),
        format="xlsx",
        size_bytes=size,
        worksheet=selected,
        worksheet_selection_policy=config.excel_worksheet_selection,
        available_worksheets=available,
    )
    dataset = ParsedDataset(metadata=metadata, headers=[])

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        workbook.close()
        if config.excel_require_non_empty_sheet:
            raise EmptyFile(
                code="empty_file",
                message="Worksheet contains no rows.",
            ) from None
        dataset.finalize_accounting()
        return dataset

    headers = [_cell_to_str(value) or "" for value in header_row]
    while headers and headers[-1] == "":
        headers.pop()
    if not any(header.strip() for header in headers):
        workbook.close()
        raise MissingHeader(code="missing_header", message="Worksheet is missing a header row.")

    for index, header in enumerate(headers):
        if header == "" and config.blank_header_cell_policy == "error":
            workbook.close()
            raise BlankHeaderCell(
                code="blank_header_cell",
                message=f"Blank header cell at column index {index}.",
                column_index=index,
            )

    duplicates = sorted({name for name in headers if headers.count(name) > 1 and name})
    if duplicates and config.duplicate_header_policy == "error":
        workbook.close()
        raise DuplicateHeader(
            code="duplicate_header",
            message=f"Duplicate header columns detected: {', '.join(duplicates)}",
            duplicate_columns=tuple(duplicates),
        )

    if len(headers) > config.max_column_count:
        workbook.close()
        raise ColumnLimitExceeded(
            code="column_limit_exceeded",
            message=f"Column count {len(headers)} exceeds limit {config.max_column_count}.",
            column_count=len(headers),
            limit=config.max_column_count,
        )

    dataset.headers = headers
    data_row_count = 0
    row_number = 1
    observed_max_column_count = len(headers)
    for raw_row in rows_iter:
        row_number += 1
        values_list = [_cell_to_str(value) for value in raw_row]
        observed_max_column_count = max(observed_max_column_count, len(values_list))
        if len(values_list) < len(headers):
            values_list.extend([None] * (len(headers) - len(values_list)))
        elif len(values_list) > len(headers):
            if config.malformed_row_policy == "reject":
                dataset.rejected_rows.append(
                    RejectedRow(
                        row_number=row_number,
                        raw_values=tuple("" if value is None else value for value in values_list),
                        reason_code="inconsistent_column_count",
                        message=(f"Expected {len(headers)} columns, found {len(values_list)}."),
                    )
                )
                dataset.issues.append(
                    ParseIssue(
                        code="inconsistent_column_count",
                        message=(
                            f"Row {row_number} has {len(values_list)} columns; "
                            f"expected {len(headers)}."
                        ),
                        severity="warning",
                        row_number=row_number,
                    )
                )
                continue
            values_list = values_list[: len(headers)]

        if _is_entirely_blank(values_list):
            if config.entirely_blank_row_policy == "reject":
                dataset.rejected_rows.append(
                    RejectedRow(
                        row_number=row_number,
                        raw_values=tuple("" if value is None else value for value in values_list),
                        reason_code="entirely_blank_row",
                        message="Row contains only blank cells.",
                    )
                )
                continue

        data_row_count += 1
        if data_row_count > config.max_row_count:
            workbook.close()
            raise RowLimitExceeded(
                code="row_limit_exceeded",
                message=f"Row count {data_row_count} exceeds limit {config.max_row_count}.",
                row_count=data_row_count,
                limit=config.max_row_count,
            )

        values = {headers[index]: values_list[index] for index in range(len(headers))}
        dataset.rows.append(ParsedRow(row_number=row_number, values=values))

    workbook.close()
    dataset.finalize_accounting()
    return dataset
